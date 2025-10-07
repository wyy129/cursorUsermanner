# {{RIPER-6:
#   Action: "Fixed & Refactored"
#   Task_ID: "#Complete-Rewrite"
#   Timestamp: "2025-10-06T23:30:00+08:00"
#   Authoring_Role: "desktop-expert"
#   Principle_Applied: "SOLID原则 - 支持订阅和用量双API接口"
#   Quality_Check: "完整的GUI应用，支持Token验证"
#   MCP_Tools_Used: ["mcp.server_time"]
# }}

import sys
import json
import requests
import urllib3
import time
from datetime import datetime
from functools import partial
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QFileDialog,
    QLabel, QMessageBox, QHeaderView, QProgressBar, QLineEdit,
    QDialog, QTextEdit, QDialogButtonBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt5.QtGui import QFont, QColor

# Excel支持
EXCEL_AVAILABLE = False
EXCEL_IMPORT_ERROR = None
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font as ExcelFont, Alignment, PatternFill
    EXCEL_AVAILABLE = True
    print("[OK] Excel support enabled (openpyxl loaded)")
except ImportError as e:
    EXCEL_IMPORT_ERROR = str(e)
    print(f"[WARN] Excel support disabled: {e}")
    print("       Tip: Run 'pip install openpyxl' to enable Excel export")

# 禁用SSL警告（因为我们禁用了SSL验证）
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


class APIWorker(QThread):
    """
    API调用工作线程
    支持订阅接口(/api/auth/stripe)和聚合使用接口(/api/dashboard/get-aggregated-usage-events)
    支持同时调用两个接口（api_type="both"）
    """
    finished = pyqtSignal(int, dict, str)  # row_index, result_data, api_type
    error = pyqtSignal(int, str, str)  # row_index, error_message, api_type
    both_finished = pyqtSignal(int, dict, dict)  # row_index, stripe_data, aggregated_data
    
    def __init__(self, row_index, token, email="", api_type="stripe", full_cookie=""):
        super().__init__()
        self.row_index = row_index
        self.token = token
        self.email = email
        self.api_type = api_type  # "stripe" 或 "aggregated" 或 "both"
        self.full_cookie = full_cookie  # 完整的Cookie字符串（可选）
        
    def run(self):
        """执行API请求 - 完全符合API文档规范"""
        try:
            # 如果是both类型，依次调用两个接口
            if self.api_type == "both":
                self._call_both_apis()
                return
            
            # 根据API类型选择URL和方法
            if self.api_type == "aggregated":
                url = "https://cursor.com/api/dashboard/get-aggregated-usage-events"
                api_name = "聚合使用接口"
                method = "POST"
            else:  # stripe
                url = "https://www.cursor.com/api/auth/stripe"
                api_name = "订阅接口"
                method = "GET"
                
            # 根据API类型设置请求头
            if self.api_type == "aggregated":
                headers = {
                    'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
                    'sec-ch-ua-mobile': '?0',
                    'sec-ch-ua-platform': '"Windows"',
                    'upgrade-insecure-requests': '1',
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
                    'accept': '*/*',
                    'sec-fetch-site': 'same-origin',
                    'sec-fetch-mode': 'cors',
                    'sec-fetch-user': '?1',
                    'sec-fetch-dest': 'empty',
                    'accept-encoding': 'gzip, deflate, br',
                    'accept-language': 'zh-CN,zh;q=0.9',
                    'content-type': 'application/json',
                    'priority': 'u=1, i',
                    'origin': 'https://cursor.com',
                    'sec-ch-ua-arch': '"x86"',
                    'sec-ch-ua-bitness': '"64"',
                    'sec-ch-ua-platform-version': '"10.0.0"',
                    'referer': 'https://cursor.com/cn/dashboard?tab=usage'
                }
            else:
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept': '*/*',
                    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8'
                }
            
            # 使用cookies字典传递session token
            cookies = {
                'WorkosCursorSessionToken': self.token
            }
            
            # 如果有完整Cookie，解析为字典格式
            if self.full_cookie:
                cookie_mode = "完整浏览器Cookie"
                # 解析完整Cookie字符串为字典
                try:
                    for cookie_str in self.full_cookie.split(';'):
                        if '=' in cookie_str:
                            key, value = cookie_str.strip().split('=', 1)
                            cookies[key] = value
                except Exception as e:
                    print(f"  [WARN] Cookie解析警告: {e}")
            else:
                cookie_mode = "仅WorkosCursorSessionToken"
            
            # 准备payload（仅用于聚合使用接口）
            payload = None
            if self.api_type == "aggregated":
                # 计算时间范围：最近30天
                end_timestamp = int(time.time() * 1000)  # 当前时间戳（毫秒）
                start_timestamp = end_timestamp - (30 * 24 * 60 * 60 * 1000)  # 30天前（毫秒）
                payload = json.dumps({
                    "teamId": -1,
                    "startDate": start_timestamp,
                    "endDate": end_timestamp
                })
            
            # 发送请求部分
            print(f"\n{'='*80}")
            print(f"[API] {api_name} 请求信息 - 账号: {self.email}")
            print(f"{'='*80}")
            print(f"[REQ] 请求URL: {url}")
            print(f"[REQ] 请求方法: {method}")
            print(f"[REQ] Cookie模式: {cookie_mode}")
            print(f"\n[COOKIE] Cookies (使用cookies参数传递):")
            for cookie_name, cookie_value in cookies.items():
                if cookie_name == 'WorkosCursorSessionToken':
                    print(f"  {cookie_name}: {cookie_value}")
                    print(f"    Token长度: {len(cookie_value)} 字符")
                    print(f"    Token格式: {'[OK] user_开头' if cookie_value.startswith('user_') else '[FAIL] 格式异常'}")
                    print(f"    Token分段: {'[OK] 包含%3A%3A' if '%3A%3A' in cookie_value else '[FAIL] 缺少分隔符'}")
                else:
                    # 其他Cookie只显示名称和长度
                    print(f"  {cookie_name}: {cookie_value[:50]}{'...' if len(cookie_value) > 50 else ''}")
            
            print(f"\n[HEADER] 请求头:")
            for key, value in headers.items():
                print(f"  {key}: {value}")
            
            if payload:
                print(f"\n[PAYLOAD] 请求体 (Payload):")
                print(f"  {payload}")
            
            print(f"\n[SEND] 发送请求 (cookies参数模式)...")
            # 根据方法类型发送请求
            ## BUG修复: 修复了致命的缩进错误。原始代码中GET请求永远不会被执行。
            if method == "POST":
                response = requests.post(url, headers=headers, cookies=cookies, data=payload, timeout=15, verify=False)
            else:
                response = requests.get(url, headers=headers, cookies=cookies, timeout=10, verify=False)
            
            print(f"\n[RESPONSE] 响应信息:")
            print(f"  状态码: {response.status_code}")
            print(f"  响应时间: {response.elapsed.total_seconds():.3f} 秒")
            print(f"  响应大小: {len(response.content)} 字节")
            print(f"  Content-Type: {response.headers.get('Content-Type', 'N/A')}")
            
            if response.status_code == 200:
                try:
                    data = response.json()
                    print(f"\n[DATA] 响应数据 (JSON格式):")
                    print(json.dumps(data, indent=2, ensure_ascii=False))
                    print(f"{'='*80}\n")
                    self.finished.emit(self.row_index, data, self.api_type)
                except json.JSONDecodeError as je:
                    print(f"\n[ERROR] JSON解析失败: {str(je)}")
                    print(f"  原始响应: {response.text[:500]}")
                    print(f"{'='*80}\n")
                    self.error.emit(self.row_index, f"JSON解析失败: {str(je)}", self.api_type)
            else:
                error_detail = ""
                try:
                    error_json = response.json()
                    error_detail = f"\n  错误详情: {json.dumps(error_json, ensure_ascii=False)}"
                except:
                    error_detail = f"\n  响应内容: {response.text[:500]}"
                
                print(f"\n[ERROR] HTTP {response.status_code}{error_detail}")
                
                # 针对不同错误码给出建议
                if response.status_code == 401:
                    print(f"\n[TIP] 建议: Token已失效，请参考'获取Token指南.md'获取新Token")
                elif response.status_code == 403:
                    print(f"\n[TIP] 建议: 访问被拒绝，可能是权限不足或账号异常")
                elif response.status_code == 429:
                    print(f"\n[TIP] 建议: 请求过于频繁，请稍后再试")
                elif response.status_code >= 500:
                    print(f"\n[TIP] 建议: 服务器错误，请稍后重试")
                    
                print(f"{'='*80}\n")
                self.error.emit(self.row_index, f"HTTP {response.status_code}", self.api_type)
                
        except requests.exceptions.Timeout:
            print(f"\n[ERROR] 请求超时: 连接超过15秒未响应")
            print(f"  URL: {url}")
            print(f"{'='*80}\n")
            self.error.emit(self.row_index, "请求超时", self.api_type)
        except requests.exceptions.ConnectionError as ce:
            print(f"\n[ERROR] 连接错误: {str(ce)}")
            print(f"  可能原因: 网络不可达、DNS解析失败、防火墙阻止")
            print(f"{'='*80}\n")
            self.error.emit(self.row_index, f"连接错误: {str(ce)}", self.api_type)
        except Exception as e:
            print(f"\n[ERROR] run方法捕获未处理异常: {type(e).__name__}")
            print(f"  错误信息: {str(e)}")
            print(f"{'='*80}\n")
            self.error.emit(self.row_index, f"{type(e).__name__}: {str(e)}", self.api_type)
    
    def _call_both_apis(self):
        """同时调用订阅接口和聚合使用接口"""
        stripe_data = None
        aggregated_data = None
        stripe_error = None
        aggregated_error = None
        
        # 先调用订阅接口
        try:
            stripe_data = self._call_single_api("stripe")
        except Exception as e:
            print(f"\n[ERROR] 订阅接口调用失败: {str(e)}")
            stripe_error = str(e)
        
        # 再调用聚合使用接口
        try:
            aggregated_data = self._call_single_api("aggregated")
        except Exception as e:
            print(f"\n[ERROR] 聚合使用接口调用失败: {str(e)}")
            aggregated_error = str(e)
        
        # 根据结果发出相应的信号
        if stripe_data and aggregated_data:
            # 两个接口都成功
            self.both_finished.emit(self.row_index, stripe_data, aggregated_data)
        elif stripe_data:
            # 只有订阅接口成功
            self.finished.emit(self.row_index, stripe_data, "stripe")
            if aggregated_error:
                self.error.emit(self.row_index, f"聚合使用接口失败: {aggregated_error}", "aggregated")
        elif aggregated_data:
            # 只有聚合使用接口成功
            self.finished.emit(self.row_index, aggregated_data, "aggregated")
            if stripe_error:
                self.error.emit(self.row_index, f"订阅接口失败: {stripe_error}", "stripe")
        else:
            # 两个接口都失败 - 必须发出error信号，否则批量查询会卡住
            error_msg = f"订阅接口: {stripe_error}; 聚合使用接口: {aggregated_error}"
            self.error.emit(self.row_index, error_msg, "both")
    
    def _call_single_api(self, api_type):
        """调用单个API并返回数据"""
        if api_type == "aggregated":
            url = "https://cursor.com/api/dashboard/get-aggregated-usage-events"
            api_name = "聚合使用接口"
            method = "POST"
        else:  # stripe
            url = "https://www.cursor.com/api/auth/stripe"
            api_name = "订阅接口"
            method = "GET"
        
        # 根据API类型设置请求头
        if api_type == "aggregated":
            headers = {
                'sec-ch-ua': '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'upgrade-insecure-requests': '1',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36',
                'accept': '*/*',
                'sec-fetch-site': 'same-origin',
                'sec-fetch-mode': 'cors',
                'sec-fetch-user': '?1',
                'sec-fetch-dest': 'empty',
                'accept-encoding': 'gzip, deflate, br',
                'accept-language': 'zh-CN,zh;q=0.9',
                'content-type': 'application/json',
                'priority': 'u=1, i',
                'origin': 'https://cursor.com',
                'sec-ch-ua-arch': '"x86"',
                'sec-ch-ua-bitness': '"64"',
                'sec-ch-ua-platform-version': '"10.0.0"',
                'referer': 'https://cursor.com/cn/dashboard?tab=usage'
            }
        else:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': '*/*',
                'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8'
            }
        
        # 使用cookies字典传递session token
        cookies = {
            'WorkosCursorSessionToken': self.token
        }
        
        # 如果有完整Cookie，解析为字典格式
        if self.full_cookie:
            try:
                for cookie_str in self.full_cookie.split(';'):
                    if '=' in cookie_str:
                        key, value = cookie_str.strip().split('=', 1)
                        cookies[key] = value
            except Exception as e:
                print(f"  [WARN] Cookie解析警告: {e}")
        
        # 准备payload（仅用于聚合使用接口）
        payload = None
        if api_type == "aggregated":
            # 计算时间范围：最近30天
            end_timestamp = int(time.time() * 1000)
            start_timestamp = end_timestamp - (30 * 24 * 60 * 60 * 1000)
            payload = json.dumps({
                "teamId": -1,
                "startDate": start_timestamp,
                "endDate": end_timestamp
            })
        
        print(f"\n{'='*80}")
        print(f"[API] {api_name} 请求信息 - 账号: {self.email}")
        print(f"{'='*80}")
        print(f"[REQ] 请求URL: {url}")
        print(f"[REQ] 请求方法: {method}")
        
        # 根据方法类型发送请求
        if method == "POST":
            response = requests.post(url, headers=headers, cookies=cookies, data=payload, timeout=15, verify=False)
        else:
            response = requests.get(url, headers=headers, cookies=cookies, timeout=10, verify=False)
        
        print(f"\n[RESPONSE] 响应信息:")
        print(f"  状态码: {response.status_code}")
        print(f"  响应时间: {response.elapsed.total_seconds():.3f} 秒")
        
        if response.status_code == 200:
            data = response.json()
            print(f"[DATA] {api_name}响应成功")
            print(f"{'='*80}\n")
            return data
        else:
            error_detail = ""
            try:
                error_json = response.json()
                error_detail = f" - {json.dumps(error_json, ensure_ascii=False)}"
            except:
                error_detail = f" - {response.text[:200]}"
            
            print(f"\n[ERROR] HTTP {response.status_code}{error_detail}")
            print(f"{'='*80}\n")
            raise Exception(f"HTTP {response.status_code}")


class CursorAccountManager(QMainWindow):
    """Cursor账号管理器主窗口"""
    
    def __init__(self):
        super().__init__()
        self.accounts_data = []
        self.filtered_accounts = []  # 用于存储过滤后的账号
        ## 优化: 初始化所有实例变量，避免使用hasattr
        self._batch_stopped = False  # 批量查询停止标志
        self._batch_querying = False # 批量查询进行中标志
        self._batch_query_type = ''
        self._batch_query_list = []
        self._batch_query_index = 0
        self._batch_completed_count = 0  # 已完成的查询数
        self._batch_max_concurrent = 10  # 最大并发数
        self._active_workers = []  # 保存活动的worker线程引用
        self.init_ui()
        
    def init_ui(self):
        """初始化用户界面"""
        excel_status = "Excel支持✓" if EXCEL_AVAILABLE else "仅JSON"
        self.setWindowTitle(f'Cursor账号管理器 v4.1 - 修复版 [{excel_status}]')
        self.setGeometry(100, 100, 1600, 800)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # 标题
        title_label = QLabel('Cursor账号信息管理系统')
        title_font = QFont('Microsoft YaHei', 18, QFont.Bold)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet('color: #2c3e50; padding: 15px;')
        main_layout.addWidget(title_label)
        
        # 控制按钮
        control_layout = QHBoxLayout()
        
        self.load_btn = QPushButton('📂 加载JSON文件')
        self.load_btn.setStyleSheet(self.get_button_style('#3498db'))
        self.load_btn.clicked.connect(self.load_json_file)
        self.load_btn.setMinimumHeight(40)
        control_layout.addWidget(self.load_btn)
        
        self.refresh_btn = QPushButton('🔄 批量更新全部')
        self.refresh_btn.setStyleSheet(self.get_button_style('#2ecc71'))
        self.refresh_btn.clicked.connect(self.batch_update_all)
        self.refresh_btn.setEnabled(False)
        self.refresh_btn.setMinimumHeight(40)
        self.refresh_btn.setToolTip('批量更新所有账号的订阅信息和使用详情（多线程并发，最大10个并发）')
        control_layout.addWidget(self.refresh_btn)
        
        self.stop_btn = QPushButton('⏹️ 停止查询')
        self.stop_btn.setStyleSheet(self.get_button_style('#e74c3c'))
        self.stop_btn.clicked.connect(self.stop_batch_query)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setMinimumHeight(40)
        self.stop_btn.setToolTip('停止正在进行的批量查询')
        control_layout.addWidget(self.stop_btn)
        
        self.export_btn = QPushButton('💾 导出数据')
        self.export_btn.setStyleSheet(self.get_button_style('#16a085'))
        self.export_btn.clicked.connect(self.export_accounts)
        self.export_btn.setEnabled(False)
        self.export_btn.setMinimumHeight(40)
        if EXCEL_AVAILABLE:
            self.export_btn.setToolTip('导出当前所有账号数据（支持Excel/JSON格式）')
        else:
            self.export_btn.setToolTip('导出当前所有账号数据（仅JSON格式，安装openpyxl后可支持Excel）')
        control_layout.addWidget(self.export_btn)
        
        self.clear_btn = QPushButton('🗑️ 清空数据')
        self.clear_btn.setStyleSheet(self.get_button_style('#95a5a6'))
        self.clear_btn.clicked.connect(self.clear_data)
        self.clear_btn.setEnabled(False)
        self.clear_btn.setMinimumHeight(40)
        control_layout.addWidget(self.clear_btn)
        
        main_layout.addLayout(control_layout)
        
        # 状态栏
        self.status_label = QLabel('请加载JSON文件开始... 提示：Token失效请查看"获取Token指南.md"')
        self.status_label.setStyleSheet('color: #7f8c8d; padding: 10px; font-size: 12px;')
        main_layout.addWidget(self.status_label)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }
            QProgressBar::chunk {
                background-color: #3498db;
            }
        """)
        main_layout.addWidget(self.progress_bar)
        
        # 搜索区域
        search_layout = QHBoxLayout()
        search_label = QLabel('🔍 搜索邮箱:')
        search_label.setStyleSheet('font-size: 14px; font-weight: bold; color: #34495e;')
        search_layout.addWidget(search_label)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText('输入邮箱地址进行搜索...')
        self.search_input.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                font-size: 13px;
                background-color: white;
            }
            QLineEdit:focus {
                border: 2px solid #3498db;
            }
        """)
        self.search_input.textChanged.connect(self.filter_accounts)
        search_layout.addWidget(self.search_input, 1)
        
        self.clear_search_btn = QPushButton('✖ 清除搜索')
        self.clear_search_btn.setStyleSheet(self.get_button_style('#95a5a6'))
        self.clear_search_btn.setMaximumWidth(120)
        self.clear_search_btn.clicked.connect(self.clear_search)
        search_layout.addWidget(self.clear_search_btn)
        
        main_layout.addLayout(search_layout)
        
        # 数据表格
        self.table = QTableWidget()
        self.setup_table()
        main_layout.addWidget(self.table)
        
        # 统计信息
        self.stats_label = QLabel('统计: 总账号数: 0 | Pro账号: 0 | 试用账号: 0 | Token有效: 0/0 | 已查询: 0')
        self.stats_label.setStyleSheet('color: #34495e; padding: 10px; font-weight: bold;')
        main_layout.addWidget(self.stats_label)
        
        # 全局样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #ecf0f1;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #bdc3c7;
                border-radius: 5px;
                gridline-color: #ecf0f1;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
        """)
        
    def setup_table(self):
        """设置表格结构"""
        headers = [
            '序号', '邮箱', '账号类型', '剩余试用天数', 
            '查询状态', '用量费用(分)', 'Token状态', '操作'
        ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        
        # 设置列宽
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 60)
        self.table.setColumnWidth(2, 120)
        self.table.setColumnWidth(3, 120)
        self.table.setColumnWidth(4, 100)
        self.table.setColumnWidth(5, 120)
        self.table.setColumnWidth(6, 120)
        self.table.setColumnWidth(7, 180)  # 容纳两个按钮
        
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        
    def get_button_style(self, color):
        """获取按钮样式"""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
            QPushButton:disabled {{
                background-color: #95a5a6;
            }}
        """
        
    def darken_color(self, hex_color):
        """使颜色变暗"""
        color = QColor(hex_color)
        h, s, v, a = color.getHsv()
        color.setHsv(h, s, int(v * 0.8), a)
        return color.name()
        
    def load_json_file(self):
        """加载JSON文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, '选择JSON文件', '', 'JSON Files (*.json);;All Files (*)'
        )
        
        if not file_path:
            return
            
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                self.accounts_data = json.load(f)
                
            if not isinstance(self.accounts_data, list):
                raise ValueError('JSON文件必须是数组格式')
                
            self.filtered_accounts = self.accounts_data.copy()
            self.display_accounts()
            self.update_statistics()
            self.refresh_btn.setEnabled(True)
            self.export_btn.setEnabled(True)
            self.clear_btn.setEnabled(True)
            self.status_label.setText(f'[OK] 成功加载 {len(self.accounts_data)} 个账号')
            self.status_label.setStyleSheet('color: #27ae60; padding: 10px; font-weight: bold;')
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'加载JSON文件失败:\n{str(e)}')
            self.status_label.setText(f'[ERROR] 加载失败: {str(e)}')
            self.status_label.setStyleSheet('color: #e74c3c; padding: 10px;')
            
    def display_accounts(self):
        """显示账号信息"""
        self.table.setRowCount(len(self.filtered_accounts))
        
        for display_idx, account in enumerate(self.filtered_accounts):
            # 找到原始索引
            original_idx = self.accounts_data.index(account)
            
            # 序号（显示原始序号）
            self.table.setItem(display_idx, 0, self.create_centered_item(str(original_idx + 1)))
            
            # 邮箱
            self.table.setItem(display_idx, 1, QTableWidgetItem(account.get('email', 'N/A')))
            
            # 账号类型
            membership_type = account.get('membershipType', 'unknown')
            membership_item = self.create_centered_item(membership_type)
            if membership_type == 'pro':
                membership_item.setBackground(QColor('#2ecc71'))
                membership_item.setForeground(QColor('white'))
            elif membership_type == 'free_trial':
                membership_item.setBackground(QColor('#f39c12'))
                membership_item.setForeground(QColor('white'))
            self.table.setItem(display_idx, 2, membership_item)
            
            # 剩余试用天数
            days_remaining = account.get('daysRemainingOnTrial')
            days_text = str(days_remaining) if days_remaining is not None else 'N/A'
            days_item = self.create_centered_item(days_text)
            if days_remaining is not None and days_remaining <= 3:
                days_item.setBackground(QColor('#e74c3c'))
                days_item.setForeground(QColor('white'))
            self.table.setItem(display_idx, 3, days_item)
            
            # 查询状态
            query_status = account.get('queryStatus', '未查询')
            query_item = self.create_centered_item(query_status)
            if query_status == '✓ 已查询':
                query_item.setBackground(QColor('#27ae60'))
                query_item.setForeground(QColor('white'))
            elif query_status == '查询中...':
                query_item.setBackground(QColor('#f39c12'))
                query_item.setForeground(QColor('white'))
            elif query_status == '✗ 失败':
                query_item.setBackground(QColor('#e74c3c'))
                query_item.setForeground(QColor('white'))
            self.table.setItem(display_idx, 4, query_item)
            
            # 用量费用
            aggregated_data = account.get('aggregatedData', {})
            total_cost = aggregated_data.get('totalCostCents', 0) if aggregated_data else 0
            cost_text = str(total_cost) if total_cost > 0 else 'N/A'
            cost_item = self.create_centered_item(cost_text)
            if total_cost > 0:
                # 根据费用高低设置不同颜色
                if total_cost > 10000:  # 超过100美元
                    cost_item.setForeground(QColor('#e74c3c'))
                    cost_item.setBackground(QColor('#ffebee'))
                elif total_cost > 5000:  # 50-100美元
                    cost_item.setForeground(QColor('#f39c12'))
                else:
                    cost_item.setForeground(QColor('#2980b9'))
            self.table.setItem(display_idx, 5, cost_item)
            
            # Token状态
            token_status = account.get('tokenStatus', '未验证')
            token_item = self.create_centered_item(token_status)
            if token_status == '✓ 有效':
                token_item.setBackground(QColor('#2ecc71'))
                token_item.setForeground(QColor('white'))
            elif token_status == '✗ 无效':
                token_item.setBackground(QColor('#e74c3c'))
                token_item.setForeground(QColor('white'))
            self.table.setItem(display_idx, 6, token_item)
            
            # 操作按钮（使用原始索引）
            button_widget = QWidget()
            button_layout = QHBoxLayout()
            button_layout.setContentsMargins(2, 2, 2, 2)
            button_layout.setSpacing(5)
            
            stripe_btn = QPushButton('📊 订阅')
            stripe_btn.setStyleSheet(self.get_button_style('#3498db'))
            stripe_btn.setMaximumWidth(85)
            stripe_btn.clicked.connect(lambda checked, i=original_idx: self.refresh_single_account(i, 'stripe'))
            button_layout.addWidget(stripe_btn)
            
            update_all_btn = QPushButton('🔄 更新全部')
            update_all_btn.setStyleSheet(self.get_button_style('#2ecc71'))
            update_all_btn.setMaximumWidth(85)
            update_all_btn.clicked.connect(lambda checked, i=original_idx: self.update_single_account(i))
            button_layout.addWidget(update_all_btn)
            
            button_widget.setLayout(button_layout)
            self.table.setCellWidget(display_idx, 7, button_widget)
            
    def create_centered_item(self, text):
        """创建居中对齐的表格项"""
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        return item
        
    def refresh_single_account(self, row_index, api_type='stripe'):
        """刷新单个账号订阅信息"""
        if row_index >= len(self.accounts_data):
            return
            
        account = self.accounts_data[row_index]
        token = account.get('auth_info', {}).get('WorkosCursorSessionToken')
        email = account.get('email', '未知')
        # 获取完整Cookie（如果有的话）
        full_cookie = account.get('auth_info', {}).get('FullCookie', '')
        
        if not token:
            QMessageBox.warning(self, '警告', f'账号 {email} 没有WorkosCursorSessionToken')
            return
        
        self.status_label.setText(f'正在查询账号 {email} 的订阅信息...')
        
        worker = APIWorker(row_index, token, email, api_type, full_cookie)
        self._active_workers.append(worker)
        
        # 使用partial避免lambda闭包问题
        worker.finished.connect(self.on_api_success)
        worker.error.connect(self.on_api_error)
        worker.finished.connect(partial(self._cleanup_worker, worker))
        worker.error.connect(partial(self._cleanup_worker, worker))
        worker.start()
    
    def update_single_account(self, row_index):
        """更新单个账号的所有信息（订阅+用量）"""
        if row_index >= len(self.accounts_data):
            return
            
        account = self.accounts_data[row_index]
        token = account.get('auth_info', {}).get('WorkosCursorSessionToken')
        email = account.get('email', '未知')
        full_cookie = account.get('auth_info', {}).get('FullCookie', '')
        
        if not token:
            QMessageBox.warning(self, '警告', f'账号 {email} 没有WorkosCursorSessionToken')
            return
        
        self.status_label.setText(f'正在更新账号 {email} 的全部信息...')
        
        worker = APIWorker(row_index, token, email, 'both', full_cookie)
        self._active_workers.append(worker)
        
        # 连接信号
        worker.both_finished.connect(self.on_both_finished)
        worker.finished.connect(self.on_api_success)
        worker.error.connect(self.on_api_error)
        worker.both_finished.connect(partial(self._cleanup_worker, worker))
        worker.finished.connect(partial(self._cleanup_worker, worker))
        worker.error.connect(partial(self._cleanup_worker, worker))
        worker.start()
        
    def show_aggregated_usage(self, row_index):
        """查询并显示聚合使用情况"""
        if row_index >= len(self.accounts_data):
            return
            
        account = self.accounts_data[row_index]
        ## BUG修复: 修复了致命的缩进错误
        token = account.get('auth_info', {}).get('WorkosCursorSessionToken')
        email = account.get('email', '未知')
        full_cookie = account.get('auth_info', {}).get('FullCookie', '')
                
        if not token:
            QMessageBox.warning(self, '警告', f'账号 {email} 没有WorkosCursorSessionToken')
            return
        
        self.status_label.setText(f'正在查询账号 {email} 的聚合使用情况（最近30天）...')
        
        worker = APIWorker(row_index, token, email, 'aggregated', full_cookie)
        self._active_workers.append(worker)
        
        # 使用partial避免lambda闭包问题
        worker.finished.connect(self.on_aggregated_success)
        worker.error.connect(self.on_api_error)
        worker.finished.connect(partial(self._cleanup_worker, worker))
        worker.error.connect(partial(self._cleanup_worker, worker))
        worker.start()
        
    def batch_update_all(self):
        """批量更新所有账号的全部信息（订阅+用量）"""
        valid_accounts = []
        
        for idx, account in enumerate(self.accounts_data):
            token = account.get('auth_info', {}).get('WorkosCursorSessionToken')
            email = account.get('email', '未知')
            full_cookie = account.get('auth_info', {}).get('FullCookie', '')
            if token:
                valid_accounts.append((idx, token, email, full_cookie))
                
        if not valid_accounts:
            QMessageBox.warning(self, '警告', '没有找到包含WorkosCursorSessionToken的账号')
            return
        
        # 设置批量查询模式
        self._batch_stopped = False
        self._batch_querying = True
        self._batch_query_type = 'both'  # 同时查询订阅和用量
        self._batch_query_list = valid_accounts
        self._batch_query_index = 0
        self._batch_completed_count = 0
        
        # 显示进度条
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(len(valid_accounts))
        self.progress_bar.setValue(0)
        
        # 禁用/启用按钮
        self.refresh_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        
        self.status_label.setText(f'开始批量更新 {len(valid_accounts)} 个账号的全部信息（最大并发数: {self._batch_max_concurrent}）...')
        self.status_label.setStyleSheet('color: #3498db; padding: 10px; font-weight: bold;')
        
        # 启动初始批次的查询（最多启动max_concurrent个）
        self._start_concurrent_queries()
            
    def on_api_success(self, row_index, data, api_type):
        """API调用成功回调（处理订阅接口）"""
        if row_index < len(self.accounts_data):
            account = self.accounts_data[row_index]
            email = account.get('email', '未知')
            
            if api_type == 'stripe':
                # 更新订阅信息
                account['membershipType'] = data.get('membershipType', 'unknown')
                account['daysRemainingOnTrial'] = data.get('daysRemainingOnTrial')
                account['tokenStatus'] = '✓ 有效'
                account['queryStatus'] = '✓ 已查询'  # 更新查询状态
            
            # 如果账号在过滤列表中，更新对应的表格行
            if account in self.filtered_accounts:
                display_idx = self.filtered_accounts.index(account)
                
                ## BUG修复: 修复了致命的缩进错误，将UI更新代码块正确地放入if语句内
                if api_type == 'stripe':
                    # 更新表格 - 订阅信息
                    membership_type = data.get('membershipType', 'unknown')
                    membership_item = self.create_centered_item(membership_type)
                    if membership_type == 'pro':
                        membership_item.setBackground(QColor('#2ecc71'))
                        membership_item.setForeground(QColor('white'))
                    elif membership_type == 'free_trial':
                        membership_item.setBackground(QColor('#f39c12'))
                        membership_item.setForeground(QColor('white'))
                    self.table.setItem(display_idx, 2, membership_item)
                    
                    days_remaining = data.get('daysRemainingOnTrial')
                    days_text = str(days_remaining) if days_remaining is not None else 'N/A'
                    days_item = self.create_centered_item(days_text)
                    if days_remaining is not None and days_remaining <= 3:
                        days_item.setBackground(QColor('#e74c3c'))
                        days_item.setForeground(QColor('white'))
                    self.table.setItem(display_idx, 3, days_item)
                    
                    # 更新Token状态
                    token_item = self.create_centered_item('✓ 有效')
                    token_item.setBackground(QColor('#2ecc71'))
                    token_item.setForeground(QColor('white'))
                    self.table.setItem(display_idx, 6, token_item)
                        
                    # 更新查询状态
                    query_item = self.create_centered_item('✓ 已查询')
                    query_item.setBackground(QColor('#27ae60'))
                    query_item.setForeground(QColor('white'))
                    self.table.setItem(display_idx, 4, query_item)
            
            # 如果不是批量查询模式，更新状态
            if not self._batch_querying:
                self.status_label.setText(f'[OK] 成功查询账号 {email} 的订阅信息')
                self.status_label.setStyleSheet('color: #27ae60; padding: 10px; font-weight: bold;')
                self.update_statistics()
        
    def on_api_error(self, row_index, error_msg, api_type):
        """API调用失败回调"""
        if row_index < len(self.accounts_data):
            account = self.accounts_data[row_index]
            email = account.get('email', '未知')
            api_name = '订阅接口' if api_type == 'stripe' else '聚合使用接口'
            print(f'查询账号 {email} 的{api_name}失败: {error_msg}')
            
            # 如果是401错误，标记Token无效
            if '401' in error_msg and api_type == 'stripe':
                account['tokenStatus'] = '✗ 无效'
                account['queryStatus'] = '✗ 失败'
                
                # 如果账号在过滤列表中，更新对应的表格行
                if account in self.filtered_accounts:
                    display_idx = self.filtered_accounts.index(account)
                    ## BUG修复: 修复了致命的缩进错误，将UI更新代码块正确地放入if语句内
                    token_item = self.create_centered_item('✗ 无效')
                    token_item.setBackground(QColor('#e74c3c'))
                    token_item.setForeground(QColor('white'))
                    self.table.setItem(display_idx, 6, token_item)
                    
                    # 更新查询状态
                    query_item = self.create_centered_item('✗ 失败')
                    query_item.setBackground(QColor('#e74c3c'))
                    query_item.setForeground(QColor('white'))
                    self.table.setItem(display_idx, 4, query_item)
            
            # 如果是聚合使用接口失败，更新查询状态
            if api_type == 'aggregated':
                account['queryStatus'] = '✗ 失败'
                
                if account in self.filtered_accounts:
                    display_idx = self.filtered_accounts.index(account)
                    query_item = self.create_centered_item('✗ 失败')
                    query_item.setBackground(QColor('#e74c3c'))
                    query_item.setForeground(QColor('white'))
                    self.table.setItem(display_idx, 4, query_item)
            
            if not self._batch_querying:
                self.update_statistics()
    
    def on_aggregated_success(self, row_index, data, api_type):
        """聚合使用接口调用成功回调"""
        if row_index < len(self.accounts_data):
            account = self.accounts_data[row_index]
            email = account.get('email', '未知')
            
            # 保存聚合数据到账号信息
            account['aggregatedData'] = data
            account['queryStatus'] = '✓ 已查询'
            
            # 如果账号在过滤列表中，更新对应的表格行
            if account in self.filtered_accounts:
                display_idx = self.filtered_accounts.index(account)
                
                # 更新查询状态
                query_item = self.create_centered_item('✓ 已查询')
                query_item.setBackground(QColor('#27ae60'))
                query_item.setForeground(QColor('white'))
                self.table.setItem(display_idx, 4, query_item)
                
                # 更新用量费用
                total_cost = data.get('totalCostCents', 0)
                cost_text = str(total_cost) if total_cost > 0 else 'N/A'
                cost_item = self.create_centered_item(cost_text)
                if total_cost > 0:
                    # 根据费用高低设置不同颜色
                    if total_cost > 10000:  # 超过100美元
                        cost_item.setForeground(QColor('#e74c3c'))
                        cost_item.setBackground(QColor('#ffebee'))
                    elif total_cost > 5000:  # 50-100美元
                        cost_item.setForeground(QColor('#f39c12'))
                    else:
                        cost_item.setForeground(QColor('#2980b9'))
                self.table.setItem(display_idx, 5, cost_item)
            
            # 如果不是批量查询模式，显示对话框
            if not self._batch_querying:
                self.show_aggregated_dialog(email, data)
                self.update_statistics()
            
            self.status_label.setText(f'[OK] 成功查询账号 {email} 的聚合使用情况')
            self.status_label.setStyleSheet('color: #27ae60; padding: 10px; font-weight: bold;')
    
    def on_both_finished(self, row_index, stripe_data, aggregated_data):
        """同时调用两个接口成功的回调"""
        if row_index < len(self.accounts_data):
            account = self.accounts_data[row_index]
            email = account.get('email', '未知')
            
            # 更新订阅信息
            account['membershipType'] = stripe_data.get('membershipType', 'unknown')
            account['daysRemainingOnTrial'] = stripe_data.get('daysRemainingOnTrial')
            account['tokenStatus'] = '✓ 有效'
            
            # 保存聚合数据
            account['aggregatedData'] = aggregated_data
            account['queryStatus'] = '✓ 已查询'
            
            # 如果账号在过滤列表中，更新对应的表格行
            if account in self.filtered_accounts:
                display_idx = self.filtered_accounts.index(account)
                
                # 更新账号类型
                membership_type = stripe_data.get('membershipType', 'unknown')
                membership_item = self.create_centered_item(membership_type)
                if membership_type == 'pro':
                    membership_item.setBackground(QColor('#2ecc71'))
                    membership_item.setForeground(QColor('white'))
                elif membership_type == 'free_trial':
                    membership_item.setBackground(QColor('#f39c12'))
                    membership_item.setForeground(QColor('white'))
                self.table.setItem(display_idx, 2, membership_item)
                
                # 更新剩余试用天数
                days_remaining = stripe_data.get('daysRemainingOnTrial')
                days_text = str(days_remaining) if days_remaining is not None else 'N/A'
                days_item = self.create_centered_item(days_text)
                if days_remaining is not None and days_remaining <= 3:
                    days_item.setBackground(QColor('#e74c3c'))
                    days_item.setForeground(QColor('white'))
                self.table.setItem(display_idx, 3, days_item)
                
                # 更新查询状态
                query_item = self.create_centered_item('✓ 已查询')
                query_item.setBackground(QColor('#27ae60'))
                query_item.setForeground(QColor('white'))
                self.table.setItem(display_idx, 4, query_item)
                
                # 更新用量费用
                total_cost = aggregated_data.get('totalCostCents', 0)
                cost_text = str(total_cost) if total_cost > 0 else 'N/A'
                cost_item = self.create_centered_item(cost_text)
                if total_cost > 0:
                    if total_cost > 10000:
                        cost_item.setForeground(QColor('#e74c3c'))
                        cost_item.setBackground(QColor('#ffebee'))
                    elif total_cost > 5000:
                        cost_item.setForeground(QColor('#f39c12'))
                    else:
                        cost_item.setForeground(QColor('#2980b9'))
                self.table.setItem(display_idx, 5, cost_item)
                
                # 更新Token状态
                token_item = self.create_centered_item('✓ 有效')
                token_item.setBackground(QColor('#2ecc71'))
                token_item.setForeground(QColor('white'))
                self.table.setItem(display_idx, 6, token_item)
            
            # 如果不是批量查询模式，更新状态
            if not self._batch_querying:
                self.status_label.setText(f'[OK] 成功更新账号 {email} 的全部信息')
                self.status_label.setStyleSheet('color: #27ae60; padding: 10px; font-weight: bold;')
                self.update_statistics()
    
    def show_aggregated_dialog(self, email, data):
        """显示聚合使用情况对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f'聚合使用详情 - {email}')
        dialog.setMinimumSize(800, 600)
        
        layout = QVBoxLayout()
        
        # 标题
        title_label = QLabel(f'账号: {email}')
        title_label.setStyleSheet('font-size: 16px; font-weight: bold; color: #2c3e50; padding: 10px;')
        layout.addWidget(title_label)
        
        # 统计信息
        summary_label = QLabel()
        summary_text = self.format_aggregated_summary(data)
        summary_label.setText(summary_text)
        summary_label.setStyleSheet('font-size: 14px; padding: 10px; background-color: #ecf0f1; border-radius: 5px;')
        layout.addWidget(summary_label)
        
        # 详细数据（JSON格式）
        detail_label = QLabel('详细数据（JSON格式）:')
        detail_label.setStyleSheet('font-weight: bold; padding: 5px;')
        layout.addWidget(detail_label)
        
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText(json.dumps(data, indent=2, ensure_ascii=False))
        text_edit.setStyleSheet('font-family: Consolas, monospace; font-size: 12px;')
        layout.addWidget(text_edit)
        
        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok)
        button_box.accepted.connect(dialog.accept)
        layout.addWidget(button_box)
        
        dialog.setLayout(layout)
        dialog.exec_()
    
    def format_aggregated_summary(self, data):
        """格式化聚合使用情况摘要"""
        summary_lines = []
        
        # 提取关键信息
        if isinstance(data, dict):
            # 总费用（最重要的信息，优先显示）
            if 'totalCostCents' in data:
                total_cost_cents = data['totalCostCents']
                total_cost_dollars = total_cost_cents / 100
                summary_lines.append(f'[COST] 总费用: {total_cost_cents} 分 (${total_cost_dollars:.2f})')
            
            # 总事件数
            total_events = len(data.get('events', []))
            summary_lines.append(f'[COUNT] 总事件数: {total_events}')
            
            # 如果有其他统计信息
            if 'totalRequests' in data:
                summary_lines.append(f'[REQUESTS] 总请求数: {data["totalRequests"]}')
            
            if 'dateRange' in data:
                date_range = data['dateRange']
                summary_lines.append(f'[DATE] 时间范围: {date_range.get("start", "N/A")} 至 {date_range.get("end", "N/A")}')
            
            # 按类型统计事件
            events = data.get('events', [])
            if events:
                event_types = {}
                for event in events:
                    event_type = event.get('type', 'unknown')
                    event_types[event_type] = event_types.get(event_type, 0) + 1
                
                summary_lines.append('\n[EVENTS] 事件类型统计:')
                for event_type, count in sorted(event_types.items(), key=lambda x: x[1], reverse=True):
                    summary_lines.append(f'  • {event_type}: {count} 次')
        
        return '\n'.join(summary_lines) if summary_lines else '暂无统计信息'
        
    def update_statistics(self):
        """更新统计信息"""
        total = len(self.accounts_data)
        pro_count = sum(1 for acc in self.accounts_data if acc.get('membershipType') == 'pro')
        trial_count = sum(1 for acc in self.accounts_data if acc.get('membershipType') == 'free_trial')
        valid_tokens = sum(1 for acc in self.accounts_data if acc.get('tokenStatus') == '✓ 有效')
        total_with_tokens = sum(1 for acc in self.accounts_data if acc.get('auth_info', {}).get('WorkosCursorSessionToken'))
        queried_count = sum(1 for acc in self.accounts_data if acc.get('queryStatus') == '✓ 已查询')
        
        # 计算总使用量
        total_usage_cents = 0
        for acc in self.accounts_data:
            aggregated_data = acc.get('aggregatedData', {})
            if aggregated_data:
                total_usage_cents += aggregated_data.get('totalCostCents', 0)
        
        # 转换为美元（保留两位小数）
        total_usage_dollars = total_usage_cents / 100
        # 格式化分（如果有小数则保留2位，否则显示整数）
        total_usage_cents_str = f'{total_usage_cents:.2f}' if total_usage_cents % 1 != 0 else f'{int(total_usage_cents)}'
        
        self.stats_label.setText(
            f'统计: 总账号数: {total} | Pro账号: {pro_count} | 试用账号: {trial_count} | '
            f'Token有效: {valid_tokens}/{total_with_tokens} | 已查询: {queried_count} | '
            f'总使用量: ${total_usage_dollars:.2f} ({total_usage_cents_str}分)'
        )
        
    def filter_accounts(self):
        """根据邮箱搜索过滤账号"""
        search_text = self.search_input.text().strip().lower()
        
        if not search_text:
            # 如果搜索框为空，显示所有账号
            self.filtered_accounts = self.accounts_data.copy()
        else:
            # 过滤出邮箱包含搜索文本的账号
            self.filtered_accounts = [
                account for account in self.accounts_data
                if search_text in account.get('email', '').lower()
            ]
        
        # 重新显示账号
        self.display_accounts()
        
        # 更新状态显示
        if search_text:
            self.status_label.setText(
                f'[SEARCH] 搜索 "{search_text}": 找到 {len(self.filtered_accounts)} 个匹配账号 (共 {len(self.accounts_data)} 个)'
            )
            self.status_label.setStyleSheet('color: #3498db; padding: 10px; font-weight: bold;')
        else:
            self.status_label.setText(f'显示全部 {len(self.accounts_data)} 个账号')
            self.status_label.setStyleSheet('color: #7f8c8d; padding: 10px;')
    
    def clear_search(self):
        """清除搜索"""
        self.search_input.clear()
        # textChanged信号会自动触发filter_accounts
    
    def _start_concurrent_queries(self):
        """启动并发查询（最多启动max_concurrent个worker）"""
        # 检查是否被停止
        if self._batch_stopped:
            # 不启动新查询，让现有的完成
            return
        
        # 计算需要启动的worker数量
        remaining = len(self._batch_query_list) - self._batch_query_index
        batch_workers = [w for w in self._active_workers if hasattr(w, '_is_batch_worker')]
        current_running = len(batch_workers)
        to_start = min(remaining, self._batch_max_concurrent - current_running)
        
        # 启动新的worker
        for _ in range(to_start):
            if self._batch_query_index >= len(self._batch_query_list):
                break
            
            # 再次检查停止标志
            if self._batch_stopped:
                break
                
            self._start_single_batch_query()
            self._batch_query_index += 1
    
    def _start_single_batch_query(self):
        """启动单个批量查询"""
        if self._batch_query_index >= len(self._batch_query_list):
            return
        
        # 获取当前要查询的账号
        idx, token, email, full_cookie = self._batch_query_list[self._batch_query_index]
        account = self.accounts_data[idx]
        
        # 更新表格显示为"查询中..."
        account['queryStatus'] = '查询中...'
        if account in self.filtered_accounts:
            display_idx = self.filtered_accounts.index(account)
            query_item = self.create_centered_item('查询中...')
            query_item.setBackground(QColor('#f39c12'))
            query_item.setForeground(QColor('white'))
            self.table.setItem(display_idx, 4, query_item)
        
        # 根据查询类型创建worker
        if self._batch_query_type == 'subscription':
            worker = APIWorker(idx, token, email, 'stripe', full_cookie)
        elif self._batch_query_type == 'usage':
            worker = APIWorker(idx, token, email, 'aggregated', full_cookie)
        else:  # both
            worker = APIWorker(idx, token, email, 'both', full_cookie)
        
        # 标记为批量worker
        worker._is_batch_worker = True
        worker._is_cleaned = False  # 添加清理标志，防止重复清理
        self._active_workers.append(worker)
        
        # 连接信号 - 根据查询类型连接不同的信号，避免重复触发
        if self._batch_query_type == 'both':
            # both模式：只连接both_finished信号
            worker.both_finished.connect(self._on_batch_query_both_success)
            worker.error.connect(self._on_batch_query_error)
            # cleanup连接到both_finished和error
            worker.both_finished.connect(partial(self._cleanup_batch_worker, worker))
            worker.error.connect(partial(self._cleanup_batch_worker, worker))
        else:
            # 单接口模式：连接finished信号
            worker.finished.connect(self._on_batch_query_success)
            worker.error.connect(self._on_batch_query_error)
            # cleanup连接到finished和error
            worker.finished.connect(partial(self._cleanup_batch_worker, worker))
            worker.error.connect(partial(self._cleanup_batch_worker, worker))
        
        worker.start()
    
    def _on_batch_query_success(self, row_index, data, api_type):
        """批量查询成功回调"""
        # 根据查询类型调用相应的处理函数
        if api_type == 'stripe':
            self.on_api_success(row_index, data, api_type)
        elif api_type == 'aggregated':
            self.on_aggregated_success(row_index, data, api_type)
        
        # 更新完成计数和进度
        self._batch_completed_count += 1
        self.progress_bar.setValue(self._batch_completed_count)
        
        # 如果已停止，不更新状态也不启动新查询
        if self._batch_stopped:
            return
        
        # 更新状态标签
        if self._batch_query_type == 'subscription':
            self.status_label.setText(
                f'批量查询订阅进行中 ({self._batch_completed_count}/{len(self._batch_query_list)}) - 并发: {len([w for w in self._active_workers if hasattr(w, "_is_batch_worker")])}'
            )
        elif self._batch_query_type == 'usage':
            self.status_label.setText(
                f'批量查询用量进行中 ({self._batch_completed_count}/{len(self._batch_query_list)}) - 并发: {len([w for w in self._active_workers if hasattr(w, "_is_batch_worker")])}'
            )
        else:
            self.status_label.setText(
                f'批量更新全部进行中 ({self._batch_completed_count}/{len(self._batch_query_list)}) - 并发: {len([w for w in self._active_workers if hasattr(w, "_is_batch_worker")])}'
            )
        
        # 启动下一个查询
        self._start_concurrent_queries()
    
    def _on_batch_query_both_success(self, row_index, stripe_data, aggregated_data):
        """批量更新全部成功回调"""
        # 调用on_both_finished处理数据更新
        self.on_both_finished(row_index, stripe_data, aggregated_data)
        
        # 更新完成计数和进度
        self._batch_completed_count += 1
        self.progress_bar.setValue(self._batch_completed_count)
        
        # 如果已停止，不更新状态也不启动新查询
        if self._batch_stopped:
            return
        
        # 更新状态标签
        self.status_label.setText(
            f'批量更新全部进行中 ({self._batch_completed_count}/{len(self._batch_query_list)}) - 并发: {len([w for w in self._active_workers if hasattr(w, "_is_batch_worker")])}'
        )
        
        # 启动下一个查询
        self._start_concurrent_queries()
    
    def _on_batch_query_error(self, row_index, error_msg, api_type):
        """批量查询错误回调"""
        # 调用原来的错误处理
        self.on_api_error(row_index, error_msg, api_type)
        
        # 更新完成计数和进度
        self._batch_completed_count += 1
        self.progress_bar.setValue(self._batch_completed_count)
        
        # 如果已停止，不更新状态也不启动新查询
        if self._batch_stopped:
            return
        
        # 更新状态标签
        if self._batch_query_type == 'subscription':
            self.status_label.setText(
                f'批量查询订阅进行中 ({self._batch_completed_count}/{len(self._batch_query_list)}) - 并发: {len([w for w in self._active_workers if hasattr(w, "_is_batch_worker")])}'
            )
        elif self._batch_query_type == 'usage':
            self.status_label.setText(
                f'批量查询用量进行中 ({self._batch_completed_count}/{len(self._batch_query_list)}) - 并发: {len([w for w in self._active_workers if hasattr(w, "_is_batch_worker")])}'
            )
        else:
            self.status_label.setText(
                f'批量更新全部进行中 ({self._batch_completed_count}/{len(self._batch_query_list)}) - 并发: {len([w for w in self._active_workers if hasattr(w, "_is_batch_worker")])}'
            )
        
        # 启动下一个查询
        self._start_concurrent_queries()
    
    def _finish_batch_query(self, message):
        """完成批量查询"""
        self._batch_querying = False
        self.progress_bar.setVisible(False)
        self.refresh_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText(message)
        
        if '完成' in message:
            self.status_label.setStyleSheet('color: #27ae60; padding: 10px; font-weight: bold;')
        else:
            self.status_label.setStyleSheet('color: #e74c3c; padding: 10px; font-weight: bold;')
        
        # 更新统计信息
        self.update_statistics()
    
    def stop_batch_query(self):
        """停止批量查询"""
        if self._batch_querying:
            self._batch_stopped = True
            self.status_label.setText('正在停止批量查询...')
            self.status_label.setStyleSheet('color: #f39c12; padding: 10px; font-weight: bold;')
            
            # 等待所有批量worker完成（异步处理避免阻塞UI）
            QTimer.singleShot(100, self._check_and_finish_stop)
        
    def export_accounts(self):
        """导出账号数据"""
        if not self.accounts_data:
            QMessageBox.warning(self, '提示', '当前没有数据可以导出')
            return
        
        # 如果Excel不可用，显示提示
        if not EXCEL_AVAILABLE:
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle('导出格式')
            msg.setText('当前仅支持JSON格式导出')
            msg.setInformativeText('提示：安装openpyxl库后可支持Excel格式导出\n运行命令: pip install openpyxl')
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            
        # 生成默认文件名（包含时间戳）
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        
        # 构建文件过滤器
        if EXCEL_AVAILABLE:
            filters = 'Excel Files (*.xlsx);;JSON Files (*.json);;All Files (*)'
            default_filename = f'cursor_accounts_export_{timestamp}.xlsx'
        else:
            filters = 'JSON Files (*.json);;All Files (*)'
            default_filename = f'cursor_accounts_export_{timestamp}.json'
        
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self, '导出账号数据', default_filename, filters
        )
        
        if not file_path:
            return
        
        try:
            # 根据文件扩展名或选择的过滤器决定导出格式
            if file_path.lower().endswith('.xlsx') and EXCEL_AVAILABLE:
                self._export_to_excel(file_path)
            elif file_path.lower().endswith('.json'):
                self._export_to_json(file_path)
            else:
                # 根据选择的过滤器决定
                if 'Excel' in selected_filter and EXCEL_AVAILABLE:
                    if not file_path.endswith('.xlsx'):
                        file_path += '.xlsx'
                    self._export_to_excel(file_path)
                else:
                    if not file_path.endswith('.json'):
                        file_path += '.json'
                    self._export_to_json(file_path)
            
            QMessageBox.information(self, '成功', f'数据已成功导出到:\n{file_path}')
            self.status_label.setText(f'[OK] 成功导出 {len(self.accounts_data)} 个账号数据')
            self.status_label.setStyleSheet('color: #27ae60; padding: 10px; font-weight: bold;')
            
        except Exception as e:
            QMessageBox.critical(self, '错误', f'导出数据失败:\n{str(e)}')
            self.status_label.setText(f'[ERROR] 导出失败: {str(e)}')
            self.status_label.setStyleSheet('color: #e74c3c; padding: 10px;')
    
    def _export_to_json(self, file_path):
        """导出为JSON格式"""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(self.accounts_data, f, ensure_ascii=False, indent=2)
    
    def _export_to_excel(self, file_path):
        """导出为Excel格式"""
        wb = Workbook()
        ws = wb.active
        ws.title = "账号数据"
        
        # 设置表头
        headers = ['序号', '邮箱', '账号类型', '剩余试用天数', '查询状态', '用量费用(分)', '用量费用(美元)', 'Token状态', 'Cookie']
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = ExcelFont(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据
        for idx, account in enumerate(self.accounts_data, 1):
            # 获取聚合数据
            aggregated_data = account.get('aggregatedData', {})
            total_cost = aggregated_data.get('totalCostCents', 0) if aggregated_data else 0
            
            ## BUG修复: 修正了导出Cookie的逻辑
            auth_info = account.get('auth_info', {})
            cookie_str = auth_info.get('FullCookie', auth_info.get('WorkosCursorSessionToken', 'N/A'))

            ## 计算美元金额
            total_cost_dollars = total_cost / 100 if total_cost > 0 else 0
            
            ## BUG修复: 修正了导出数据时使用的错误键名
            row_data = [
                idx,
                account.get('email', 'N/A'),
                account.get('membershipType', 'N/A'),
                account.get('daysRemainingOnTrial', 'N/A'),
                account.get('queryStatus', '未查询'),
                total_cost if total_cost > 0 else 'N/A',
                f'${total_cost_dollars:.2f}' if total_cost > 0 else 'N/A',
                account.get('tokenStatus', '未验证'),
                cookie_str
            ]
            ws.append(row_data)
            
            # 设置数据行样式
            row_num = idx + 1
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='center' if col_num != 9 else 'left', vertical='center')
                
                # 根据内容设置颜色
                if col_num == 5:  # 查询状态
                    query_status = account.get('queryStatus', '未查询')
                    if query_status == '✓ 已查询':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = ExcelFont(color="006100")
                    elif query_status == '✗ 失败':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = ExcelFont(color="9C0006")
                elif col_num == 6 or col_num == 7:  # 用量费用(分)和用量费用(美元)
                    if total_cost > 10000:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = ExcelFont(color="9C0006")
                    elif total_cost > 5000:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = ExcelFont(color="9C5700")
                elif col_num == 8:  # Token状态
                    token_status = account.get('tokenStatus', '未验证')
                    if token_status == '✓ 有效':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = ExcelFont(color="006100")
                    elif token_status == '✗ 无效':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = ExcelFont(color="9C0006")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 80
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        wb.save(file_path)
        
    def clear_data(self):
        """清空数据"""
        reply = QMessageBox.question(
            self, '确认', '确定要清空所有数据吗?',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.accounts_data = []
            self.filtered_accounts = []
            self.table.setRowCount(0)
            self.refresh_btn.setEnabled(False)
            self.export_btn.setEnabled(False)
            self.clear_btn.setEnabled(False)
            self.search_input.clear()
            self.status_label.setText('数据已清空，请重新加载JSON文件')
            self.status_label.setStyleSheet('color: #7f8c8d; padding: 10px;')
            self.update_statistics()
    
    def _cleanup_worker(self, worker):
        """清理完成的worker线程（单个查询）"""
        try:
            if worker in self._active_workers:
                self._active_workers.remove(worker)
            # 延迟删除worker对象
            QTimer.singleShot(100, worker.deleteLater)
        except Exception as e:
            print(f"[WARN] Worker cleanup error: {e}")
    
    def _cleanup_batch_worker(self, worker):
        """清理完成的批量查询worker线程"""
        try:
            # 防止重复清理
            if hasattr(worker, '_is_cleaned') and worker._is_cleaned:
                return
            
            # 标记为已清理
            worker._is_cleaned = True
            
            # 从活动列表中移除
            if worker in self._active_workers:
                self._active_workers.remove(worker)
            
            # 延迟删除worker对象
            QTimer.singleShot(100, worker.deleteLater)
            
            # 检查是否被停止
            if self._batch_stopped:
                # 检查是否还有运行中的批量worker
                batch_workers = [w for w in self._active_workers if hasattr(w, '_is_batch_worker') and w.isRunning()]
                if not batch_workers:
                    self._finish_batch_query('[STOPPED] 批量查询已停止')
            # 检查是否所有批量查询都已完成
            elif self._batch_querying and self._batch_completed_count >= len(self._batch_query_list):
                # 确保没有正在运行的批量worker
                batch_workers = [w for w in self._active_workers if hasattr(w, '_is_batch_worker') and w.isRunning()]
                if not batch_workers:
                    self._finish_batch_query('[SUCCESS] 批量查询完成！')
        except Exception as e:
            print(f"[WARN] Batch worker cleanup error: {e}")
    
    def _check_and_finish_stop(self):
        """检查并完成停止操作"""
        if not self._batch_stopped:
            return
        
        # 检查是否还有运行中的批量worker
        batch_workers = [w for w in self._active_workers if hasattr(w, '_is_batch_worker') and w.isRunning()]
        
        if batch_workers:
            # 还有worker在运行，继续等待
            self.status_label.setText(f'正在停止批量查询...（等待{len(batch_workers)}个线程完成）')
            QTimer.singleShot(100, self._check_and_finish_stop)
        else:
            # 所有worker都已完成，完成停止
            if self._batch_querying:
                self._finish_batch_query('[STOPPED] 批量查询已停止')
    
    def closeEvent(self, event):
        """窗口关闭事件 - 清理所有活动的线程"""
        # 停止批量查询
        if self._batch_querying:
            self._batch_stopped = True
        
        # 等待所有活动的worker线程完成
        if self._active_workers:
            print(f"\n[INFO] 等待 {len(self._active_workers)} 个线程完成...")
            for worker in self._active_workers[:]:  # 使用副本遍历
                if worker.isRunning():
                    worker.wait(1000)  # 最多等待1秒
                    if worker.isRunning():
                        worker.terminate()  # 强制终止
                        worker.wait()
        
        event.accept()


def main():
    """主程序入口"""
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    window = CursorAccountManager()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
